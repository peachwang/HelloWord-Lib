# -*- coding: utf-8 -*-  
import sys, os; sys.path.extend([os.path.abspath(_[0]) for _ in os.walk(os.path.join(os.getcwd(), '../'))]);
from util import *
from openpyxl import Workbook, load_workbook

class Table :

    def __init__(self) :
        self.tableData         = []
        self.fieldNames        = []
        self.primaryFieldNames = None

    def importFromTxt(self, filename, cast = None, seperator = '\t') :
        fin = open(filename, 'r')
        self.fieldNames = fin.readline().strip('\n').split(seperator)
        tableData = load_txt(fin, fields = self.fieldNames, cast = cast, sep = seperator)
        for rowData in tableData :
            self.tableData.append(_TableRow(rowData))
        return self

    def importFromXlsx(self, filename, sheetName = None, cast = None) :
        wb = load_workbook(filename)
        ws = wb.active if sheetName == None else wb.get_sheet_by_name(sheetName)
        rows = list(ws.rows)
        self.fieldNames = [str(cell.value) for cell in list(rows[0])] # must be str
        for row in rows[1:] :
            rowData = {}
            for fieldIndex, cell in enumerate(row) :
                fieldName = self.fieldNames[fieldIndex]
                rowData[fieldName] = cell.value
                if cast is not None :
                    rowData[fieldName] = cast[fieldName](rowData[fieldName])
            self.tableData.append(_TableRow(rowData))
        return self

    def setPrimaryField(self, primaryFieldNames) :
        self.primaryFieldNames = primaryFieldNames
        return self

    def setByRules(self, rules) :
        for rule in rules :
            for rowIndex, tableRow in enumerate(self.tableData) :
                args = [rowIndex, tableRow, self]
                criterion  = True if rule.get('criterion') is None else rule['criterion'](*args)
                if criterion :
                    fieldName  = rule['fieldName'](*args)
                    fieldValue = rule['fieldValue'](*(args + [fieldName]))
                    tableRow.setField(fieldName, fieldValue)

    def __get__(self, primaryFieldValueRanges, defaultValue = None) :
        if self.primaryFieldNames is None : return None
        else : 
            indexes = self.calculateIndexes(self.primaryFieldNames, fieldValues)

    def getData(self) :
        return [tableRow.getData() for tableRow in self.tableData]

    def getFieldNames(self) :
        return self.fieldNames.copy()

    def getFieldValues(self, fieldNames = None, defaultValue = None, criterion = None) :
        if fieldNames == None : fieldNames = self.getFieldNames()
        fieldValues = []
        for rowIndex, tableRow in enumerate(self.tableData) :
            if criterion is None or criterion(rowIndex, tableRow) :
                fieldValues.append(tableRow.getFieldValues(fieldNames, defaultValue))
        return fieldValues

    # 
    # @param str|list fieldNames
    #   <fieldName str>
    # @return list
    #   <index int>
    def calculateRowIndexes(self, fieldNames, fieldValueRanges) :
        rowIndexes = []
        if type(fieldNames) is str : fieldNames = [fieldNames]
        for rowIndex, tableRow in enumerate(self.tableData) :
            fieldValues = tableRow.getFieldValues(fieldNames)
            isMatch = True
            for fieldIndex, fieldValue in enumerate(fieldValues) :
                if (type(fieldValueRanges[fieldIndex]) is list \
                and fieldValue not in fieldValueRanges[fieldIndex]) \
                or (fieldValue != fieldValueRanges[fieldIndex]) :
                    isMatch = False
                    break
            if isMatch : rowIndexes.append(rowIndex)
        return rowIndexes

    def indexMatch(self, otherTable, targetFieldName, targetFieldNameOther, lookupFieldNameList, lookupFieldNameListOther) :
        for tableRow in self.tableData :
            pass

    def exportToTxt(self, filename, seperator = '\t', defaultValue = '') :
        fout = open(filename, 'w')
        fieldNames = self.getFieldNames()
        lines = [seperator.join(fieldNames)]
        for tableRow in self.tableData :
            lines.append(seperator.join(map(str, tableRow.getFieldValues(fieldNames, defaultValue))))
        fout.write('\n'.join(lines) + '\n')
        fout.flush()

    def exportToXlsx(self, filename, defaultValue = '') :
        wb = Workbook()
        ws = wb.active
        fieldNames = self.getFieldNames()
        ws.append(fieldNames)
        for tableRow in self.tableData :
            ws.append(tableRow.getFieldValues(fieldNames, defaultValue))
        wb.save(filename)

class _TableRow :
    def __init__(self, rowData = {}) :
        self.rowData = rowData

    def __get__(self, fieldName, defaultValue = None) :
        return self.rowData.get(fieldName, defaultValue)

    def __set__(self, fieldName, fieldValue) :
        self.rowData[fieldName] = fieldValue

    def getData(self) :
        return self.rowData.copy()

    def getFieldNames(self) :
        return self.rowData.keys()

    # 
    # @param str|list fieldNames
    # @return scalar|list
    #   <fieldValue scalar>
    def getFieldValues(self, fieldNames = None, defaultValue = None) :
        if fieldNames is None : fieldNames = self.getFieldNames()
        if type(fieldNames) is str :
            return self.__get__(fieldNames, defaultValue)
        return [self.__get__(fieldName, defaultValue) for fieldName in fieldNames]

    def setField(self, fieldName, fieldValue) :
        if self.__get__(fieldName) == fieldValue :
            return 0
        else :
            self.__set__(fieldName, fieldValue)
            return 1

    def setFields(self, namesToValues) :
        return sum([self.setField(fieldName, fieldValue) for fieldName, fieldValue in namesToValues.items()])

